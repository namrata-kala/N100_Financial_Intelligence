import pandas as pd

from src.screener.engine import ScreenerEngine
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def generate_peer_report():

    engine = ScreenerEngine()

    df = engine.load_data()
    df = engine.calculate_composite_score(df)

    writer = pd.ExcelWriter(
        "output/peer_comparison.xlsx",
        engine="openpyxl"
    )

    report_columns = [
        "Sector Rank",
        "Company",
        "Quality Score",
        "ROE (%)",
        "Revenue CAGR (%)",
        "PAT CAGR (%)",
        "Debt/Equity",
        "Free Cash Flow (Cr)",
    ]

    for sector in sorted(df["broad_sector"].dropna().unique()):

        sector_df = (
            df[df["broad_sector"] == sector]
            .dropna(subset=["composite_quality_score"])
            .copy()
        )

        sector_df["Sector Rank"] = (
            sector_df["composite_quality_score"]
            .rank(method="dense", ascending=False)
            .astype(int)
        )

        sector_df = sector_df.sort_values("Sector Rank")

        sector_df["composite_quality_score"] = (
            sector_df["composite_quality_score"].round(2)
        )

        sector_df["debt_to_equity"] = (
            sector_df["debt_to_equity"].round(2)
        )

        sector_df["return_on_equity_pct"] = (
            sector_df["return_on_equity_pct"].round(2)
        )

        sector_df["revenue_cagr_5yr"] = (
            sector_df["revenue_cagr_5yr"].round(2)
        )

        sector_df["pat_cagr_5yr"] = (
            sector_df["pat_cagr_5yr"].round(2)
        )

        sector_df["free_cash_flow_cr"] = (
            sector_df["free_cash_flow_cr"].round(2)
        )

        sector_df = sector_df.rename(columns={
            "company_id": "Company",
            "composite_quality_score": "Quality Score",
            "return_on_equity_pct": "ROE (%)",
            "revenue_cagr_5yr": "Revenue CAGR (%)",
            "pat_cagr_5yr": "PAT CAGR (%)",
            "debt_to_equity": "Debt/Equity",
            "free_cash_flow_cr": "Free Cash Flow (Cr)",
        })

        sheet = sector[:31]

        sector_df[report_columns].to_excel(
            writer,
            sheet_name=sheet,
            index=False
        )

    workbook = writer.book

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        header_fill = PatternFill(
            fill_type="solid",
            start_color="4F81BD"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        gold_fill = PatternFill(
            fill_type="solid",
            start_color="FFD966"
        )

        for cell in ws[2]:
            cell.fill = gold_fill

    workbook = writer.book

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        for column in ws.columns:

            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 3


    writer.close()

    print("peer_comparison.xlsx created successfully!")


if __name__ == "__main__":
    generate_peer_report()